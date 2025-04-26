import pygame
from openpyxl import *

wb=load_workbook('quest.xlsx')
sheet=wb.active

pygame.init()
class settings():
	def __init__(self):
		self.bg_game=pygame.image.load\
		("images/bg1.jpg")	#картинка заднего фона
		self.icon=pygame.image.load\
		("images/icon.jpg") #иконка игры
		self.bg_menu=pygame.image.load\
		("images/bg2.jpg") #картинка меню
		self.menu_butt1=pygame.image.load\
		("images/menu_butt1.png") #картинка неактивной кнопки меню
		self.menu_butt2=pygame.image.load\
		("images/menu_butt2.png") #картинка активной кнопки меню
		self.redbutton=pygame.image.load\
		("images/redbutt.png")	#картинка пройденного вопроса
		self.quest_win=pygame.image.load\
		("images/question1.png")	#окно с вопросом
		self.AnswerTrueFalse1=pygame.image.load\
		("images/AnswerTrue.png") #картинка правильного ответа
		self.AnswerTrueFalse2=pygame.image.load\
		("images/AnswerFalse.png") #картинка неправильного ответа
		self.nameteams=pygame.image.load\
		("images/nameteams.png")
		self.check_exit=False
		self.question_act=False	#переменная работы окна с вопросом
		self.end_game=False #переменная конца игры
		
		self.delay=0
		self.Answer=None
		self.names=None
		
		self.WINwidth=800	#ширина окна
		self.WINheight=600	#высота окна
		self.caption="СВОЯ ИГРА"	#название игры
		self.fps=20	#кол-во кадров в секунду
		
		#размеры кнопок с вопросами
		self.butt1W=100
		self.butt1H=82
		
		#размеры кнопок пропуска вопроса и выхода в меню после игры
		self.butt2W=120
		self.butt2H=22
		
		#размеры кнопок меню
		self.butt3W=221
		self.butt3H=48
		
		#названия шрифтов
		self.font1='hatten.ttf'
		self.font2='arial'
		self.font3='19847.otf'
		
		#кегль шрифтов
		self.pin=25
		self.pin1=25
		self.pin2=72
		self.pin3=12
		self.pin4=40
		
		#шрифты
		self.font=[
		pygame.font.SysFont\
		(self.font2,self.pin), #шрифт для текста вопросов
		pygame.font.Font\
		(self.font1,self.pin1), #шрифт названий тем и кол-ва очков
		pygame.font.SysFont\
		(self.font2,self.pin1), #
		pygame.font.Font\
		(self.font3,self.pin2), #шрифт для надписи "Своя игра" и кол-ва очков после конца игры
		pygame.font.SysFont\
		(self.font2,self.pin3,bold=True), #шрифт маленького текста на кнопках
		pygame.font.SysFont\
		(self.font2,self.pin1,bold=True) #шрифт для информации
		]
		
		#массив картинок неактивных кнопок
		self.image_buttons=[pygame.image.load\
		("images/butt1.png"),
		pygame.image.load("images/butt2.png"),
		pygame.image.load("images/butt3.png"),
		pygame.image.load("images/butt4.png"),
		pygame.image.load("images/butt5.png")]
		
		#массив картинок активных кнопок
		self.click_buttons=[pygame.image.load\
		("images/butt11.png"),
		pygame.image.load("images/butt21.png"),
		pygame.image.load("images/butt31.png"),
		pygame.image.load("images/butt41.png"),
		pygame.image.load("images/butt51.png")]
		
		#массив картинок для кнопок "пропустить(+0)", "проверить", "меню" после конца игры, "меню" для выхода из инструкции
		self.im_sys_buttons=[pygame.image.load\
		("images/sysbutt1.png"),
		pygame.image.load("images/sysbutt11.png")]
		
		self.menu_num=-1 #изначальный номер кнопки меню
		self.quest_num=-1 #изначальный номер вопроса
		self.score=[0,0] #кол-во очков
		self.butt_count=25 #кол-во кнопок с вопросами
		
		self.Next=0 #переменная, показывающая ту команду, которая ходит
		self.musicPlay=True #переменная, показывающая, можно ли проигрывать музыку или нет
		self.time_out=[int(sheet.cell(row=2,column=8)\
		.value)+1,int(sheet.cell(row=2,column=8)\
		.value)+1] #кол-во секунд для таймера отсчёта
