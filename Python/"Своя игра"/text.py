import pygame
from settings import *
from openpyxl import *

_set=settings()

wb=load_workbook('quest.xlsx') #загрузка рабочей книги Excel
sheet=wb.active #установка первого листа как активного

#Все названия тем, инструкция, вопросы и ответы к ним читаются из листа Excel


#массив названий тем
quest_theme=[]
for i in range(5):
	quest_theme.append\
	(sheet.cell(row=(i*5)+2,column=4).value)

#массив вопросов	
questions=[]
for i in range(2,27):
	#считываем из каждой ячейки во 2 столбике вопросы
	questions.append\
	(sheet.cell(row=i,column=2).value) 

#массив ответов на вопросы
answer=[]
for i in range(2,27):
	#считываем из каждой ячейки в 3 столбике ответы на вопросы
	answer.append(str(\
	sheet.cell(row=i,column=3).value))

#функция отрисовки названий тем
def draw_theme(sc):
	for i in range(len(quest_theme)):
		text=_set.font[1].render\
		(quest_theme[i],1,(255,255,255))
		sc.blit(text,(580,65+(i*100)))

#функция отрисовки вопросов
def draw_quest(sc,i):
	text=questions[i]
	
	dt=(_set.pin//3)+3
	xt=140
	i,kf,kl=0,0,0
	
	strings=[]
	
	while i<len(text):	
		if xt<=680:
			if text[i]==' ':
				i+=1
				kl=i
				xt+=dt
			elif text[i]!=' ':
				if i+1==len(text):
					kl=i+1
					strings.append\
					(text[kf:kl])
					break
				else:
					i+=1
					xt+=dt
		else:
			strings.append\
			(text[kf:kl])
			kf=i=kl
			xt=140
	
	yt,xt=140,140

	for i in range(len(strings)):
		sc.blit(_set.font[0].render\
		(strings[i], 0, (0,0,0)), \
		(xt, yt+(i*_set.pin)))
	
#функция отрисовки очков команд внизу экрана и кто сейчас ходит
def draw_score(sc,team,score,Next):
	
	text=_set.font[1].render\
	(str(team[0])+": %s"%score[0],\
	1,(255,0,0))
	sc.blit(text,(30,567))
	
	text=_set.font[1].render\
	(str(team[1])+": %s"%score[1],\
	1,(0,0,255))
	sc.blit(text,(390,567))
	
	if Next==0:
		text=_set.font[1].render\
		('Ходит команда: %r'%team[Next],\
		1,(255,0,0))
	else:
		text=_set.font[1].render\
		('Ходит команда: %r'%team[Next],\
		1,(0,0,255))
	sc.blit(text,(235,-4))

#функция отрисовки очков посередине экрана после конца вопросов
def end_game(sc,team,score):
	if score[0]>score[1]:
		sc.blit(_set.font[5].render\
		("Победила команда: %s"%team[0],\
		1,(255, 109, 105)),(180,140))
	elif score[0]<score[1]:
		sc.blit(_set.font[5].render\
		("Победила команда: %s"%team[1],\
		1,(255, 109, 105)),(180,140))
	else:
		sc.blit(_set.font[5].render\
		("НИЧЬЯ!",1,(255, 109, 105)),\
		(350,140))
	sc.blit(_set.font[5].render\
	("Команда "+str(team[0])+" заработала: %s очков"\
	%score[0],1,(255,0,0)),(130,250))
	sc.blit(_set.font[5].render\
	("Команда "+str(team[1])+" заработала: %s очков"\
	%score[1],1,(0,0,255)),(130,300))
	
#функция отрисовывает предупреждение о выходе из игры
def draw_exit(sc):
	text=str(sheet.cell(row=2,column=6).value)
	dt=(_set.pin1//3)+3
	xt=140
	i,kf,kl=0,0,0
	strings=[]
	
	while i<len(text):	
		if xt<=640:
			if text[i]==' ':
				i+=1
				kl=i
				xt+=dt
			elif text[i]!=' ':
				if i+1==len(text):
					kl=i+1
					strings.append\
					(text[kf:kl])
					break
				else:
					i+=1
					xt+=dt
		else:
			strings.append(text[kf:kl])
			kf=i=kl
			xt=140
	
	yt,xt=120,140
	
	for i in range(len(strings)):
		sc.blit(_set.font[5].render\
		(strings[i], 0, (0,0,0)),\
		(xt, yt+(i*_set.pin1)))
		
		

#надписи на кнопках "пропустить", "проверить", "меню", "уверен", "не уверен"
#def draw_textbutt(sc,i):
#	if i==0:
#		sc.blit(_set.font[4].render\
#		("ПРОПУСТИТЬ(+0)",1,(0,0,0)),(426,425))
#	elif i==1:
#		sc.blit(_set.font[4].render\
#		('ПРОВЕРИТЬ',1,(0,0,0)),(280,425))
#	elif i==2:
#		sc.blit(_set.font[4].render\
#		('ВЫХОД',1,(0,0,0)),(697,570))
#	elif i==3:
#		sc.blit(_set.font[4].render\
#		('НЕ УВЕРЕН',1,(0,0,0)),(454,425))
#	elif i==4:
#		sc.blit(_set.font[4].render\
#		('УВЕРЕН',1,(0,0,0)),(263,425))
#	elif i==5:
#		sc.blit(_set.font[4].render\
#		("МЕНЮ",1,(0,0,0)),(377,405))

text_butt=["ПРОПУСТИТЬ(+0)","ПРОВЕРИТЬ",\
"ВЫХОД","НЕ УВЕРЕН","УВЕРЕН","МЕНЮ"]
pos_text=[(423,427),(275,427),(697,570),\
(454,425),(263,425),(377,405)]
#надписи на кнопках "пропустить", "проверить", "меню", "уверен", "не уверен"
def draw_textbutt(sc,i):
	sc.blit(_set.font[4].render\
	(text_butt[i],1,(0,0,0)),pos_text[i])

#функция отрисовки теста инструкции к игре
def draw_info(sc):
	text=str(sheet.cell(row=2,column=5).value)
	dt=(_set.pin1//3)+3
	xt=40
	i=0
	kf=0
	kl=0
	strings=[]
	
	while i<len(text):	
		if xt<=640:
			if text[i]==' ':
				i+=1
				kl=i
				xt+=dt
			elif text[i]!=' ':
				if i+1==len(text):
					kl=i+1
					strings.append\
					(text[kf:kl])
					break
				else:
					i+=1
					xt+=dt
		else:
			strings.append(text[kf:kl])
			kf=i=kl
			xt=40
	
	yt=60
	xt=40
	for i in range(len(strings)):
		sc.blit(_set.font[5].render\
		(strings[i], 0, (255, 239, 89)),\
		(xt, yt+(i*_set.pin1)))
